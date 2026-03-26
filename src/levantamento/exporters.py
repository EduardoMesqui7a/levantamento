from __future__ import annotations

import io
import json
from dataclasses import asdict, is_dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PALETA = {
    "azul_escuro": "102A43",
    "verde": "0F766E",
    "verde_claro": "D1FAE5",
    "areia": "F8FAFC",
    "cinza": "E2E8F0",
    "texto": "0F172A",
    "branco": "FFFFFF",
    "amarelo": "FEF3C7",
}


def _flatten_eap(items, nivel: int = 0):
    rows = []
    for item in items or []:
        rows.append(
            {
                "item": item.get("item", ""),
                "descricao": item.get("descricao", ""),
                "unidade": item.get("unidade", ""),
                "quantidade": item.get("quantidade", ""),
                "preco_unitario": float(item.get("preco_unitario", 0.0) or 0.0),
                "preco_total": float(item.get("preco_total", 0.0) or 0.0),
                "nivel": nivel,
                "observacoes": item.get("observacoes", ""),
            }
        )
        rows.extend(_flatten_eap(item.get("filhos", []), nivel + 1))
    return rows


def _flatten_materiais(items):
    rows = []
    for item in items or []:
        rows.append(
            {
                "descricao": item.get("descricao", ""),
                "unidade": item.get("unidade", ""),
                "quantidade": item.get("quantidade", ""),
                "origem": item.get("origem", ""),
                "confianca": float(item.get("confianca", 0.0) or 0.0),
                "categoria": item.get("categoria", ""),
            }
        )
    return rows


def _flatten_audit(audit: dict[str, object] | None) -> list[dict[str, object]]:
    if not audit:
        return []

    rows: list[dict[str, object]] = []
    for sinal in audit.get("sinais_confirmados", []) or []:
        rows.append({"tipo": "Sinal confirmado", "descricao": str(sinal), "motivo": "", "categoria": ""})
    for criterio in audit.get("criterios_aceitacao", []) or []:
        rows.append({"tipo": "Critério", "descricao": str(criterio), "motivo": "", "categoria": ""})
    for item in audit.get("itens_rejeitados", []) or []:
        rows.append(
            {
                "tipo": "Rejeitado",
                "descricao": str(item.get("descricao", "")),
                "motivo": str(item.get("motivo", "")),
                "categoria": str(item.get("categoria", "")),
            }
        )
    rows.append(
        {
            "tipo": "Confiança global",
            "descricao": f"{float(audit.get('confianca_global', 0.0) or 0.0):.0%}",
            "motivo": str(audit.get("observacoes_finais", "")),
            "categoria": "",
        }
    )
    return rows


def result_to_json(result) -> str:
    if is_dataclass(result):
        result = asdict(result)
    return json.dumps(result, ensure_ascii=False, indent=2)


class DataFrameExports:
    def __init__(self, result):
        self.result = result

    def eap_dataframe(self) -> pd.DataFrame:
        df = pd.DataFrame(_flatten_eap(self.result.get("eap", [])))
        if df.empty:
            return pd.DataFrame(columns=["ITEM", "DESCRIÇÃO", "UNIDADE", "QUANTIDADE", "PREÇO UNITÁRIO", "PREÇO TOTAL"])
        return df.rename(
            columns={
                "item": "ITEM",
                "descricao": "DESCRIÇÃO",
                "unidade": "UNIDADE",
                "quantidade": "QUANTIDADE",
                "preco_unitario": "PREÇO UNITÁRIO",
                "preco_total": "PREÇO TOTAL",
            }
        )[["ITEM", "DESCRIÇÃO", "UNIDADE", "QUANTIDADE", "PREÇO UNITÁRIO", "PREÇO TOTAL"]]

    def materiais_dataframe(self) -> pd.DataFrame:
        df = pd.DataFrame(_flatten_materiais(self.result.get("materiais", [])))
        if df.empty:
            return pd.DataFrame(columns=["DESCRIÇÃO", "UNIDADE", "QUANTIDADE", "ORIGEM", "CONFIANÇA", "CATEGORIA"])
        return df.rename(
            columns={
                "descricao": "DESCRIÇÃO",
                "unidade": "UNIDADE",
                "quantidade": "QUANTIDADE",
                "origem": "ORIGEM",
                "confianca": "CONFIANÇA",
                "categoria": "CATEGORIA",
            }
        )

    def _merge_title_block(self, ws, end_col: int, end_row: int = 4) -> None:
        for row in range(1, end_row + 1):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
            cell = ws.cell(row=row, column=1)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=PALETA["verde_claro"])
            cell.border = Border(
                left=Side(style="thin", color=PALETA["cinza"]),
                right=Side(style="thin", color=PALETA["cinza"]),
                top=Side(style="thin", color=PALETA["cinza"]),
                bottom=Side(style="thin", color=PALETA["cinza"]),
            )
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 34
        ws.row_dimensions[4].height = 24

    def _write_title_block(self, ws, end_col: int) -> None:
        self._merge_title_block(ws, end_col)
        ws["A1"] = "EAP ORÇAMENTÁRIA"
        ws["A1"].font = Font(size=18, bold=True, color=PALETA["verde"])
        ws["A2"] = f"Projeto: {self.result.get('tipo_projeto', 'Não identificado')}"
        ws["A2"].font = Font(size=11, color=PALETA["texto"])
        ws["A3"] = f"Resumo: {self.result.get('resumo', '')}"
        ws["A3"].font = Font(size=10, color=PALETA["texto"])
        ws["A4"] = f"Total de itens da EAP: {self.result.get('total_itens_eap', 0)}"
        ws["A4"].font = Font(size=10, color=PALETA["texto"])

    def _style_sheet(self, ws, header_row: int, freeze_cell: str) -> None:
        thin = Side(style="thin", color=PALETA["cinza"])
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor=PALETA["verde"])
        title_fill = PatternFill("solid", fgColor=PALETA["verde_claro"])

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = Font(color=PALETA["branco"], bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = freeze_cell
        ws.sheet_view.showGridLines = False

        for row in range(1, 5):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = title_fill

    def _set_widths(self, ws, widths: dict[str, float]) -> None:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def _style_budget_rows(self, ws, rows, start_row: int) -> None:
        fills = {
            0: PatternFill("solid", fgColor="EFF6FF"),
            1: PatternFill("solid", fgColor="F0FDFA"),
            2: PatternFill("solid", fgColor="FFFBEB"),
            3: PatternFill("solid", fgColor="F8FAFC"),
        }
        currency = '"R$" #,##0.00'
        thin = Side(style="thin", color=PALETA["cinza"])
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for offset, row_data in enumerate(rows, start=start_row):
            nivel = int(row_data.get("nivel", 0) or 0)
            fill = fills.get(min(nivel, 3), fills[3])
            for col in range(1, 7):
                cell = ws.cell(row=offset, column=col)
                cell.fill = fill
                cell.border = border
                if col == 2:
                    indent = max(nivel, 0) * 2
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=indent)
                elif col in (1, 3):
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

            ws.cell(row=offset, column=4).number_format = "General"
            ws.cell(row=offset, column=5).number_format = currency
            ws.cell(row=offset, column=6).number_format = currency

    def to_excel_bytes(self) -> bytes:
        wb = Workbook()
        ws_orc = wb.active
        ws_orc.title = "Orçamento"

        self._write_title_block(ws_orc, end_col=6)

        headers = ["ITEM", "DESCRIÇÃO", "UNIDADE", "QUANTIDADE", "PREÇO UNITÁRIO", "PREÇO TOTAL"]
        header_row = 6
        for col, header in enumerate(headers, start=1):
            cell = ws_orc.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color=PALETA["branco"])
            cell.fill = PatternFill("solid", fgColor=PALETA["azul_escuro"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        eap_rows = _flatten_eap(self.result.get("eap", []))
        for row_idx, row in enumerate(eap_rows, start=header_row + 1):
            ws_orc.cell(row=row_idx, column=1, value=row["item"])
            ws_orc.cell(row=row_idx, column=2, value=row["descricao"])
            ws_orc.cell(row=row_idx, column=3, value=row["unidade"])
            ws_orc.cell(row=row_idx, column=4, value=row["quantidade"])
            ws_orc.cell(row=row_idx, column=5, value=row["preco_unitario"])
            ws_orc.cell(row=row_idx, column=6, value=row["preco_total"])

        self._style_budget_rows(ws_orc, eap_rows, start_row=header_row + 1)
        self._style_sheet(ws_orc, header_row=header_row, freeze_cell="A7")
        self._set_widths(
            ws_orc,
            {
                "A": 12,
                "B": 58,
                "C": 14,
                "D": 14,
                "E": 18,
                "F": 18,
            },
        )

        ws_materiais = wb.create_sheet("Materiais")
        self._merge_title_block(ws_materiais, end_col=6)
        ws_materiais["A1"] = "LISTA DE MATERIAIS"
        ws_materiais["A1"].font = Font(size=18, bold=True, color=PALETA["verde"])
        ws_materiais["A2"] = f"Projeto: {self.result.get('tipo_projeto', 'Não identificado')}"
        ws_materiais["A2"].font = Font(size=11, color=PALETA["texto"])
        ws_materiais["A3"] = "Somente itens que a IA classificou como materiais, componentes ou serviços técnicos pertinentes."
        ws_materiais["A3"].font = Font(size=10, color=PALETA["texto"])
        ws_materiais["A4"] = "Linhas de observação e recomendações gerais são filtradas na etapa de revisão."
        ws_materiais["A4"].font = Font(size=10, color=PALETA["texto"])

        material_headers = ["DESCRIÇÃO", "UNIDADE", "QUANTIDADE", "ORIGEM", "CONFIANÇA", "CATEGORIA"]
        for col, header in enumerate(material_headers, start=1):
            cell = ws_materiais.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color=PALETA["branco"])
            cell.fill = PatternFill("solid", fgColor=PALETA["azul_escuro"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(_flatten_materiais(self.result.get("materiais", [])), start=7):
            ws_materiais.cell(row=row_idx, column=1, value=row["descricao"])
            ws_materiais.cell(row=row_idx, column=2, value=row["unidade"])
            ws_materiais.cell(row=row_idx, column=3, value=row["quantidade"])
            ws_materiais.cell(row=row_idx, column=4, value=row["origem"])
            ws_materiais.cell(row=row_idx, column=5, value=row["confianca"])
            ws_materiais.cell(row=row_idx, column=6, value=row["categoria"])

        self._style_sheet(ws_materiais, header_row=6, freeze_cell="A7")
        for cell in ws_materiais[6]:
            cell.fill = PatternFill("solid", fgColor=PALETA["verde"])
        for row in range(7, ws_materiais.max_row + 1):
            ws_materiais.cell(row=row, column=5).number_format = '0.0%'
        self._set_widths(
            ws_materiais,
            {
                "A": 56,
                "B": 12,
                "C": 12,
                "D": 34,
                "E": 12,
                "F": 18,
            },
        )

        ws_resumo = wb.create_sheet("Resumo")
        ws_resumo["A1"] = "RESUMO DO LEVANTAMENTO"
        ws_resumo["A1"].font = Font(size=18, bold=True, color=PALETA["verde"])
        ws_resumo.merge_cells("A1:B1")
        ws_resumo["A2"] = f"Projeto: {self.result.get('tipo_projeto', 'Não identificado')}"
        ws_resumo["A2"].font = Font(size=11, color=PALETA["texto"])
        ws_resumo.merge_cells("A2:B2")
        for row in (1, 2):
            ws_resumo.cell(row=row, column=1).fill = PatternFill("solid", fgColor=PALETA["verde_claro"])
            ws_resumo.cell(row=row, column=1).border = Border(
                left=Side(style="thin", color=PALETA["cinza"]),
                right=Side(style="thin", color=PALETA["cinza"]),
                top=Side(style="thin", color=PALETA["cinza"]),
                bottom=Side(style="thin", color=PALETA["cinza"]),
            )
        ws_resumo.row_dimensions[1].height = 26
        ws_resumo.row_dimensions[2].height = 22
        ws_resumo.row_dimensions[3].height = 34
        ws_resumo.row_dimensions[4].height = 24
        resumo_linhas = [
            ("Resumo", self.result.get("resumo", "")),
            ("Avisos", " | ".join(self.result.get("avisos", [])) or "Nenhum"),
            ("Arquivo de origem", self.result.get("metadados", {}).get("arquivo", "")),
            ("Páginas processadas", self.result.get("metadados", {}).get("paginas", 0)),
            ("Páginas com OCR", self.result.get("metadados", {}).get("paginas_ocr", 0)),
            ("Total de itens da EAP", self.result.get("total_itens_eap", 0)),
        ]
        for row_idx, (label, value) in enumerate(resumo_linhas, start=3):
            ws_resumo.cell(row=row_idx, column=1, value=label)
            ws_resumo.cell(row=row_idx, column=2, value=value)
            ws_resumo.cell(row=row_idx, column=1).font = Font(bold=True, color=PALETA["texto"])
            ws_resumo.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
        self._set_widths(ws_resumo, {"A": 22, "B": 84})

        ws_auditoria = wb.create_sheet("Auditoria")
        self._merge_title_block(ws_auditoria, end_col=4)
        ws_auditoria["A1"] = "AUDITORIA TÉCNICA"
        ws_auditoria["A1"].font = Font(size=18, bold=True, color=PALETA["verde"])
        ws_auditoria["A2"] = f"Projeto: {self.result.get('tipo_projeto', 'Não identificado')}"
        ws_auditoria["A2"].font = Font(size=11, color=PALETA["texto"])
        ws_auditoria["A3"] = "Revisão final dos sinais técnicos, critérios de aceitação e itens descartados pela IA."
        ws_auditoria["A3"].font = Font(size=10, color=PALETA["texto"])
        ws_auditoria["A4"] = "A aba serve para transparência da análise e rastreabilidade do raciocínio técnico."
        ws_auditoria["A4"].font = Font(size=10, color=PALETA["texto"])

        audit_headers = ["TIPO", "DESCRIÇÃO", "MOTIVO", "CATEGORIA"]
        for col, header in enumerate(audit_headers, start=1):
            cell = ws_auditoria.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color=PALETA["branco"])
            cell.fill = PatternFill("solid", fgColor=PALETA["azul_escuro"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        audit_rows = _flatten_audit(self.result.get("auditoria", {}))
        for row_idx, row in enumerate(audit_rows, start=7):
            ws_auditoria.cell(row=row_idx, column=1, value=row["tipo"])
            ws_auditoria.cell(row=row_idx, column=2, value=row["descricao"])
            ws_auditoria.cell(row=row_idx, column=3, value=row["motivo"])
            ws_auditoria.cell(row=row_idx, column=4, value=row["categoria"])

        self._style_sheet(ws_auditoria, header_row=6, freeze_cell="A7")
        self._set_widths(
            ws_auditoria,
            {
                "A": 18,
                "B": 56,
                "C": 56,
                "D": 18,
            },
        )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def dataframe_exports(result) -> DataFrameExports:
    return DataFrameExports(result)
